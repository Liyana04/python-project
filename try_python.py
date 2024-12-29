# -*- coding: utf-8 -*-
"""try python.ipynb

Automatically generated by Colab.

Original file is located at
    https://colab.research.google.com/drive/1vcxeF97iBg2xMEJC26jIMwM3oNYQqYkP
"""

'''from openpyxl import Workbook
from openpyxl.utils import get_column_letter
wb=Workbook()
destination_filename='new_workbook.xlsx'
ws=wb.active
ws.title="New Active Workbook"

for row in range(1,10):
  ws.append(range(100))

wssheet=wb.create_sheet(title="Number Two")
wssheet['A5']='some value here'
wssheet['A5']
ws['C3'].value


another_worksheet=wb.create_sheet(title="Third Sheet")
for row in range(1,5):
  for col in range(2,10):
    _ =another_worksheet.cell(column=col,
                            row=row,
                             value="{0}".format(get_column_letter(col)))
print(another_worksheet['B1'].value)

from openpyxl import load_workbook
from google.colab import drive
drive.mount("/content/gdrive")


wb=load_workbook(filename =
                 '/content/gdrive/My Drive/'  +'new_workbook.xlsx')
sheet=wb['New Active Workbook']

sheet.merge_celLs("C2:E3")
sheet.unmerge_celLs("C2:E3")

sheet.merge_cells(start_row=4)
wb.save('/content/gdrive/My Drive/'  +'new_workbook.xlsx')'''

'''from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from google.colab import drive
drive.mount("/content/gdrive")

wb=Workbook()
destination_filename='new_workbook_hello.xlsx'

sheet=wb.active
sheet.title="Table Worksheet"
data=[
    ["leopard",9000,400,4],
    ["Tabby cat",90,48,4],
    ["wolf",1200,444,4],
    ["tiger",1400,90,4],
    ["cats",9000000,23,4],
]

sheet.append(["Animal","Ate","Weight","Paws"])

for row in data:
  sheet.append(row)

tableObject=Table(displayName="AnimalsTable",
                  ref="A1:D6")
table_style=TableStyleInfo(name="TableStyleMedium10",
                           showRowStripes=False,
                           showColumnStripes=True,
                           showFirstColumn=True)
tableObject.tableStyleInfo=table_style
sheet.add_table(tableObject)
wb.save("/content/gdrive/My Drive/"+ destination_filename)'''


#another project for create xlsx file in gdrive
'''from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from google.colab import drive
drive.mount("/content/gdrive")

wb=Workbook()
destination_filename='yana_animal.xlsx'

sheet=wb.active
sheet.title="Table Worksheet"
data=[
    ["leopard",1,0,0,1,0,4,1],
    ["bird",0,1,0,0,1,2,0],
    ["Tabby cat",1,0,0,1,0,4,1],
    ["eagle",1,0,0,1,0,4,0],
    ["fish",0,0,1,1,0,0,0],
    ["wolf",1,0,0,1,0,4,1],
    ["tiger",1,0,0,1,0,4,0],
    ["worm",0,0,0,1,0,8,1],
    ["cats",1,0,0,1,0,4,1],
]

sheet.append(["Name","Fur","Wings","Gills","Teeth","Beak","Number of Legs","Fangs"])

for row in data:
  sheet.append(row)

tableObject=Table(displayName="AnimalsTable",
                  ref="A1:I9")


wb.save("/content/gdrive/My Drive/"+ destination_filename)'''

#another project
'''import Pandas as pd

from google.colab import drive
drive.mount("/content/gdrive")

dataset = pd.read_csv("/content/gdrive/My Drive/" + "animals.csv")'''

#another project
'''from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import NamedStyle, Font, Border, Side
#from openpyxl.worksheet.table import Table, TableStyleInfo
from google.colab import drive
drive.mount("/content/gdrive")

wb=Workbook()
destination_filename="new_format_cells.xlsx"

sheet=wb.active
sheet.title="Format Cells"

#a1Cell=sheet["A1"]
#fontStyle=Font(italic=True)
#a1Cell.font=fontStyle
#a1Cell.font=Font(color="AAA000", italic=True)

newNamedStyle=NamedStyle(name="newNamedStyle")
newNamedStyle.font=Font(bold=True, size=40)

newBorder=Side(style="thin", color="BBB111")
newNamedStyle.border=Border(top=newBorder,
                            bottom=newBorder,
                            left=newBorder,
                            right=newBorder)
sheet["B1"].style=newNamedStyle
sheet["C1"].style="newNamedStyle"

wb.save("/content/gdrive/My Drive/"+ destination_filename)'''

#another project
'''from openpyxl import Workbook
from openpyxl.styles import Font
from copy import copy
from google.colab import drive
drive.mount("/content/gdrive")

wb=Workbook()
destination_filename="new_format_cells.xlsx"

sheet=wb.active
sheet.title="Format Cells"

originalFont=Font(name="Helvetica", size=30)
copyFont=copy(originalFont)
copyFont.size=20

a1Cell=sheet["A1"]
a2Cell=sheet["A2"]

a1Cell.font=originalFont
a2Cell.font=copyFont

wb.save("/content/gdrive/My Drive/"+ destination_filename)'''

#new project/chart projects start here
'''from openpyxl import Workbook
from openpyxl.chart import Reference, Series, ScatterChart, AreaChart, BarChart, LineChart, PieChart'''
#from copy import copy
'''from google.colab import drive
drive.mount("/content/gdrive")

wb=Workbook()
destination_filename="chart_workbook.xlsx"
("/content/gdrive/My Drive/"+ destination_filename
sheet=wb.active
sheet.title="Chart Worksheet"'''

#bar chart starts here
'''
for index in range(20):
  sheet.append([index])

chart_val=Reference(sheet,
                    min_col=1,
                    min_row=1,
                    max_col=1,
                    max_row=20)

chartObj=BarChart()
chartObj.add_data(chart_val)

#add chart to sheet, BARCHART
sheet.add_chart(chartObj,"C1")'''

#line charts start here
'''rows=[
    ["Employee Number","Years Employed","Salary"],
    [1,2,10000],[3,2,12000],[1,12,102000],[7,9,67000],[10,0.5,22000]
]

for row in rows:
  sheet.append(row)

lineChartObject=LineChart()
chart_vals=Reference(sheet,
                     min_row=1,
                     min_col=1,
                     max_row=5,
                     max_col=3)

lineChartObject.add_data(chart_vals)
sheet.add_chart(lineChartObject,"F1")

lineChartObject.title="Employees"
lineChartObject.x_axis.title="X"
lineChartObject.y_axis.title="Y"

chart_style=lineChartObject.series[0]#to declare
chart_style.graphicalProperties.line.solidFill="HHHHHH"
chart_style.graphicalProperties.line.width=10
chart_style.graphicalProperties.dashStyle="sysDot"'''

#pie chart starts here
'''pie_val=[
    ["Products","Sales"],
    ["Amazing Ninja Game",5000000],
    ["Little Nightmares",120000],
    ["Temple Run",921000]
]

pieChartObj=PieChart()

for row in pie_val:
  sheet.append(row)

labels_reference=Reference(sheet,
                           min_row=2,
                           min_col=1,
                           max_row=4,
                           max_col=1)
val_ref=Reference(sheet,
                  min_row=2,
                  min_col=2,
                  max_row=4,
                  max_col=2)

pieChartObj.add_data(val_ref)
pieChartObj.set_categories(labels_reference)


sheet.add_chart(pieChartObj,"F1")'''

#area chart starts here
'''chart_values=[
    ["Employee Number","Sales","Hours"],
    [1,100000,300],
    [2,350000,256],
    [3,600000,112],
    [4,200000,564],
]

for row in chart_values:
  sheet.append(row)

areaChartObj=AreaChart()

#max_row count include header
labels_reference=Reference(sheet,
                           min_row=1,
                           min_col=1,
                           max_row=5)

values_reference=Reference(sheet,
                           min_row=1,
                           min_col=2,
                           max_row=5,
                           max_col=3)

areaChartObj.add_data(values_reference)
areaChartObj.set_categories(labels_reference)

sheet.add_chart(areaChartObj, "E1")'''

#scatter chart starts here
'''chart_values=[
    ["Employee Number","Hours Worked","Type"],
    [1,100, 123],
    [2,200, 234],
    [3,150, 345]
]

for row in chart_values:
  sheet.append(row)

scatterChartObject=ScatterChart()
values_reference=Reference(sheet,
                           min_row=2,
                           min_col=1,
                           max_row=4)

for index in range(2,4):
  values=Reference(sheet,
                   min_row=1,
                   min_col=index,
                   max_row=4)

seriesObject=Series(values,
                    values_reference)
scatterChartObject.series.append(seriesObject)

sheet.add_chart(scatterChartObject, "F1")'''


#wb.save("/content/gdrive/My Drive/"+ destination_filename)

#new project for datetime
'''from openpyxl import Workbook
from google.colab import drive
drive.mount("/content/gdrive")
from datetime import date
from datetime import timedelta
import calendar

wb=Workbook()
destination_filename="employee_timelog.xlsx"

sheet=wb.active
sheet.title="employee timelog"

sheet["A1"]="Day of the week"
sheet["B1"]="Date"
sheet["C1"]="Starting Time"
sheet["D1"]="Ending Time"'''

'''
#something not working
for each_column in sheet.columns:
  max_length = 0
  column_name = each_column[0].column
  for each_cell in each_column:
    try:
      if len(str(each_cell.value)) > max_length:
        max_length=len(each_cell.value)
    except:
      pass
  new_width = max_length * 2
  sheet.column_dimensions[column_name].width = new_width'''

#CAPS MEASN CONSTANT
'''STARTING_DATE_CELL     = "B2"
ENDING_DATE_CELL       ="B16"
TODAY                  =date.today()
daysToSubtract         =14

for row in sheet[STARTING_DATE_CELL:ENDING_DATE_CELL]:
  for cell in row:
    dateValue=TODAY-timedelta(days=daysToSubtract)
    daysToSubtract-=1
    sheet[cell.coordinate]=dateValue

STARTING_DAY_CELL    ="A2"
ENDING_DAY_CELL      ="A16"

for row in sheet[STARTING_DAY_CELL:ENDING_DAY_CELL]:
  for cell in row:
    dateValue=TODAY-timedelta(days=daysToSubtract)
    daysToSubtract-=1 #14 days before today
    sheet[cell.coordinate]=calendar.day_name[dateValue.weekday()-1]

wb.save("/content/gdrive/My Drive/"+ destination_filename)'''

#new project for export to txt
'''
#create the customer_information.xlsx first
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from google.colab import drive
drive.mount("/content/gdrive")

wb=Workbook()
destination_filename='customer_information.xlsx'

sheet=wb.active
sheet.title="Table Worksheet"
data=[
    ["Leo","Johor","A4 Paper",10],
    ["Sakinah","Johor","Notebook",3],
    ["Lee","Kelantan","Pencil",12],
    ["Resmond","Kedah","Ruler",4],
    ["Ali","Sarawak","A4 Paper",5],
    ["Aliyah","Sarawak","A4 Paper",45],
    ["Beyonce","Terengganu","Notebook",19],
    ["Ray","Kedah","Scissor",2]
]

sheet.append(["Name","State","Bought","Amount"])

for row in data:
  sheet.append(row)

tableObject=Table(displayName="AnimalsTable",
                  ref="A1:I9")


wb.save("/content/gdrive/My Drive/"+ destination_filename)'''

'''
#create txt file rom xlsx
import openpyxl
from google.colab import drive
drive.mount("/content/gdrive")
import pprint

destination_filename = "customer_information.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook("/content/gdrive/My Drive/" + destination_filename)
sheet = wb["Customers"]
customer_data = {}

# Loop through the rows to collect data
for row in range(2, sheet.max_row + 1):
    name = sheet["A" + str(row)].value
    state = sheet["B" + str(row)].value
    bought = sheet["C" + str(row)].value
    amount = sheet["D" + str(row)].value

    # Populate customer_data dictionary
    customer_data.setdefault(state, {})
    customer_data[state]['name'] = name
    customer_data[state]['bought'] = bought
    customer_data[state]['amount'] = amount

# Write the customer data to a text file
text_file = open("/content/gdrive/My Drive/" + "textFile.txt", "w")
text_file.write(pprint.pformat(customer_data))
text_file.close()

print("Data has been written to textFile.txt successfully.")'''

#another project to update information
'''
import openpyxl
from google.colab import drive
drive.mount("/content/gdrive")
import pprint

destination_filename = "customer_information.xlsx"

# Load workbook and sheet
wb = openpyxl.load_workbook("/content/gdrive/My Drive/" + destination_filename)
sheet = wb["Customers"]

CUSTOMER_TO_UPDATE={
    "Leo":("Kelantan", "Notebook", 11),
    "Beyonce":9
}

for row in range(2, 10):
    customer = sheet.cell(row=row, column=1).value  # Get the customer's name
    if customer in CUSTOMER_TO_UPDATE:
        # Update the respective columns for the customer
        if customer == "Leo":
            amount, bought, state = CUSTOMER_TO_UPDATE[customer]
            sheet.cell(row=row, column=2).value = amount  # Update amount
            sheet.cell(row=row, column=3).value = bought  # Update bought
            sheet.cell(row=row, column=4).value = state   # Update state

wb.save("/content/gdrive/My Drive/"+ destination_filename)'''

#other new project to build stock chart
'''import openpyxl
from openpyxl.chart import Reference, StockChart
from google.colab import drive
drive.mount("/content/gdrive")
import pprint

destination_filename = "stock_chart.xlsx"

# Load workbook and sheet
wb = Workbook()
sheet=wb.active
sheet.title="Stock Chart"

chart_values=[
    ["Date","Volume","Start","Min","Max","End"],
    ["2020-09-09",10000,10,11,9,10.5],
    ["2020-09-10",11000,8,12,7.9,8.1],
    ["2020-09-11",12000,9,13,8.5,10],
]

for row in chart_values:
  sheet.append(row)
  stockChartObject=StockChart()
  labels_reference=Reference(sheet,
                             min_row=2,
                             min_col=1,
                             max_row=4)
values_reference=Reference(sheet,
                           min_row=1,
                           min_col=4,
                           max_row=4,
                           max_col=6)

stockChartObject.add_data(values_reference)
stockChartObject.set_categories(labels_reference)

sheet.add_chart(stockChartObject, "B11")
wb.save("/content/gdrive/My Drive/"+ destination_filename)'''

#other new project to build doughnut chart
'''import openpyxl
from openpyxl.chart import Reference, DoughnutChart
from google.colab import drive
drive.mount("/content/gdrive")
import pprint

destination_filename = "doughnut_chart.xlsx"

# Load workbook and sheet
wb = Workbook()
sheet=wb.active
sheet.title="Stock Chart"

chart_values=[
    ["Product","Q1","Q2"],
    ["Pluvo",10000,1100],
    ["Rens",2000,2100],
    ["Escape Code",3200,3000]
]

for row in chart_values:
  sheet.append(row)
  doughnutChartObject=DoughnutChart()
  labels_reference=Reference(sheet,
                             min_row=2,
                             min_col=1,
                             max_row=4)
values_reference=Reference(sheet,
                           min_row=1,
                           min_col=2,
                           max_row=4)

doughnutChartObject.add_data(values_reference)
doughnutChartObject.set_categories(labels_reference)

sheet.add_chart(doughnutChartObject, "D1")
wb.save("/content/gdrive/My Drive/"+ destination_filename)'''


#another project about web scraping
# Install necessary packages
!apt-get update  # Update package list
!apt-get install -y chromium-chromedriver  # Install Chromium and Chromedriver
!pip install selenium  # Install Selenium

from selenium import webdriver
import time
import bs4  # BeautifulSoup library
import openpyxl
from openpyxl import Workbook
from google.colab import drive
drive.mount("/content/gdrive")

# Set up Chrome options for headless browsing
options = webdriver.ChromeOptions()
options.add_argument("--headless")  # Run Chrome in headless mode
options.add_argument("--no-sandbox")  # Disable the sandboxing
options.add_argument("--disable-dev-shm-usage")  # Overcome limited resource problems

# Set the path for the ChromeDriver


# Create a new Chrome browser instance with specified options
browser = webdriver.Chrome(options=options)

# Open the target YouTube channel's videos page
browser.get("https://www.youtube.com/@xodiac.official/videos")

# Scroll down to load more content
page_length = browser.execute_script("window.scrollTo(0, document.body.scrollHeight); return document.body.scrollHeight;")
match = False
while not match:
    last = page_length
    time.sleep(2)  # Pause for 2 seconds
    page_length = browser.execute_script("window.scrollTo(0, document.body.scrollHeight); return document.body.scrollHeight;")
    if last == page_length:
        match = True

# Parse the page source with BeautifulSoup
soup = bs4.BeautifulSoup(browser.page_source, "html.parser")
anchor_tags = soup.findAll("a")
titles_list = [anchor.get("title") for anchor in anchor_tags if anchor.get("title") is not None]

# Close the browser
browser.quit()

# Load or create a new workbook and select the active sheet
destination_filename = "/content/gdrive/My Drive/web_scrapper.xlsx"
try:
    wb = openpyxl.load_workbook(destination_filename)
except FileNotFoundError:
    wb = Workbook()

sheet = wb.active

# Write titles to the Excel sheet
row_count = 1
for title in titles_list:
    current_cell = sheet.cell(row=row_count, column=1)
    current_cell.value = title
    row_count += 1

# Save the workbook
wb.save(destination_filename)

print("Titles have been successfully written to web_scrapper.xlsx.")