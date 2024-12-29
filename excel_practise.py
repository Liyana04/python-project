import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# def process_workbook(filename) can use function
wb = xl.load_workbook('Book1.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
# return exact value as above
cell = sheet.cell(1, 1)

for row in range(1, sheet.max_row + 1):
    print(row)


for item in range(2, sheet.max_row + 1):
    # 3 here means 3rd column
    cell = sheet.cell(item, 3)
    print(cell.value)
    # print('After corrected price')
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(item, 4)
    corrected_price_cell.value = corrected_price

values = Reference(sheet,
          min_row=2,
          max_row=sheet.max_row,
          min_col=4,
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')
wb.save('Book2.xlsx')